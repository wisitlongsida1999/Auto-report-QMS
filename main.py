
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select #For Dropdown
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec

import pandas as pd
import openpyxl
import os
# import xlsxwriter
import shutil
import time
import re
import datetime
import win32com.client

from config import config

# # [1] Scraping data From QIT ##
# cisco_mail="npanichc@cisco.com"
# cisco_pass="white2$$$$QW"
# cisco_mail = "anuchma@cisco.com"
# cisco_pass = "Nhoinhoi022653#"
cisco_mail = config['username']
cisco_pass = config['password']


driver=webdriver.Chrome()
driver.get("https://www-plmprd.cisco.com/Agile/")


time.sleep(1)
WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//input[@id="userInput"]'))).send_keys(cisco_mail)
time.sleep(1)
WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//input[@name="login-button"]'))).click()
time.sleep(1)



time.sleep(1)
WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//input[@id="passwordInput"]'))).send_keys(cisco_pass)
time.sleep(1)
WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//input[@id="login-button"]'))).click()


time.sleep(5)


two_fa_url=driver.current_url
time.sleep(1)
while(two_fa_url==driver.current_url):
    pass

print("Login to QMS is success!!!")


time.sleep(1)
driver.get("https://www-plmprd.cisco.com/Agile/")



# 18-oct-2021
#check old download in download folder 
# user = "\Wisit Longsida"  ## Config ###
user = "\wisitl"  ## Config ###


qms_original_folder  = r"C:\Users" + user + "\Downloads" + "\SearchResults.xls"


if os.path.exists(qms_original_folder):
    os.remove(qms_original_folder)
    print("Already Removed old SearchResults.xls in download folder")
else:
    print("The SearchResults.xls file does not exist in download folder")




WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//td[@id="ygtvt3"]//div[@class="ygtvspacer"]'))).click()


WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//a[@title="QIT_40G&100G"]'))).click()  

WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//span[@id="More_110span"]'))).click()

WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//li[@id="yui-gen159"]'))).click()

#Check file is downloaded   #Current path from download: C:\Users\Wisit Longsida



qms_downloaded=False
while(not qms_downloaded):

    if(os.path.exists(qms_original_folder )):
        qms_downloaded=True
        print("Download Search Result is success!!!")
    else:
        print("Wait for download QMS file")
        time.sleep(1)






# [2] Move SearchResults File to same folder


if os.path.exists("SearchResults.xls"):
    os.remove("SearchResults.xls")
    print("Already Removed old SearchResults.xls")
else:
    print("The SearchResults.xls file does not exist")



target_folder = os.getcwd()
shutil.move(qms_original_folder , target_folder)
print("From", qms_original_folder , ">>>>>>>>To>>>>>>>>", os.getcwd(), "is success!!!")



# [3] Download data from Cesium

date_start_search = "01-01-2021"
test_area = "fpcbpm,fpcbpm2,fpcbdg"




time.sleep(1)
#Get Serial from xlsx
wb = pd.read_excel("SearchResults.xls")
serial_search_cs = wb["Site Received Serial Number (Affected Items)"]
joined_string = ",".join(serial_search_cs)
time.sleep(1)
joined_string = str(joined_string)
print(joined_string)
print("Number of all serials :", int(len(re.findall(",", joined_string)) + 1))

driver.maximize_window()
time.sleep(1)
driver.get("https://cesium.cisco.com/apps/PolarisSearch/AdvanceSearch")


# 18-oct-2021
cs_original_folder  = r"C:\Users" + user + "\Downloads" + "\download.xlsx"


if os.path.exists(cs_original_folder):
    os.remove(cs_original_folder)
    print("Already Removed old download.xlsx in download folder")
else:
    print("The download.xlsx file does not exist in download folder")


# Set up starting date time to search
search_page_render=False
while(not search_page_render):
    try:
        WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//input[@class="md-datepicker-input"]'))).send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
        driver.find_elements_by_xpath('//input[@class="md-datepicker-input"]')[0].send_keys(date_start_search)
        search_page_render=True
    except:
        print("Waiting for Seach page load...")
        driver.refresh()
        time.sleep(1)








driver.find_elements_by_xpath('//div[@class="form-group col-xs-12 col-md-6"]')[0].find_elements_by_xpath('//textarea')

driver.find_elements_by_xpath('//textarea')[0].send_keys(joined_string)
time.sleep(1)
driver.find_elements_by_xpath('//textarea')[3].send_keys(test_area)
time.sleep(1)
driver.find_element_by_xpath('//button[@class="btn btn-primary"]').click()
time.sleep(1)
print("Search Succesfull !!!")





try:
    WebDriverWait(driver, 30).until(ec.visibility_of_element_located((By.XPATH, '//div[@class="ui-grid-icon-container"]'))).click()  # Dropdown
except:
    print("!!!!!!!!!!! Missing Dropdown !!!!!!!!!!!!!")
    exit()
time.sleep(1)

driver.find_elements_by_xpath('//button[@class="ui-grid-menu-item ng-binding"]')[2].click()


print("STEP 3 >>> Download Test Result of All Serials is Complete!!!!")

time.sleep(1)

driver.quit()











# [4] Check file is downloaded   #Current path from download: C:\Users\Wisit Longsida

cs_downloaded=False
while(not cs_downloaded):

    if(os.path.exists(cs_original_folder )):
        cs_downloaded=True
        print("Download Cesium test result is success!!!")
    else:
        print("Wait for download Cesium file")
        time.sleep(1)

if os.path.exists("download.xlsx"):
    os.remove("download.xlsx")
    print("Already Removed old download.xlsx")
else:
    print("The download.xlsx file does not exist")

shutil.move(cs_original_folder , target_folder)
print("From", cs_original_folder , ">>>>>>>>To>>>>>>>>", os.getcwd(), "is success!!!")








# Error xlsx file from cesium can not open  // Can solved

def fix_cs_excel_format(path,name):
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    workbook = excel.Workbooks.Open("{}/{}".format(path,name))

    excel.DisplayAlerts = False
    workbook.SaveAs("{}/{}".format(path,name))
    print("Fix Cesium excel format DONE !")
    excel.Quit()


def xls_to_xlsx(path,name):
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    workbook = excel.Workbooks.Open("{}/{}".format(path,name))

    excel.DisplayAlerts = False
    workbook.SaveAs("{}/{}x".format(path,name), FileFormat = 51)
    print("XLS to XLSX format DONE !")
    excel.Quit()




# Copy FA CASE & SN To sheet2
xls_to_xlsx(target_folder,'SearchResults.xls')

time.sleep(1)
wb1 = openpyxl.load_workbook('SearchResults.xlsx')
ws1 = wb1.active

wb2 = openpyxl.load_workbook('TEMPLATE_40G&100G.xlsx')
ws2_2 = wb2["Sheet2"]

for cell in ws1['C:C']:
    ws2_2.cell(row=cell.row, column=1, value=cell.value)

for cell in ws1['B:B']:
    ws2_2.cell(row=cell.row, column=2, value=cell.value)




#Delete Search XLSX
if os.path.exists("SearchResults.xlsx"):
    os.remove("SearchResults.xlsx")
    print("Already Removed old SearchResults.xlsx")
else:
    print("The SearchResults.xlsx file does not exist")


#########################
#Create Review file from template  https://stackoverflow.com/questions/48620532/copying-an-entire-column-using-openpyxl-in-python-3



fix_cs_excel_format(target_folder,'download.xlsx')

time.sleep(1)



wb3 = openpyxl.load_workbook('download.xlsx')
ws3 = wb3.active


ws2_1 = wb2["Sheet1"]
count_cell=0
for cell in ws3['A:A']:
    ws2_1.cell(row=cell.row, column=1, value=cell.value)
    count_cell+=1

for cell in ws3['B:B']:
    ws2_1.cell(row=cell.row, column=3, value=cell.value)

for cell in ws3['E:E']:
    ws2_1.cell(row=cell.row, column=4, value=cell.value)

for cell in ws3['G:G']:
    ws2_1.cell(row=cell.row, column=5, value=cell.value)

#     Date Name of file
date=datetime.datetime.now().strftime("%x")
date_lst=''
for i in date:
    if i != '/':
        date_lst+=i
    else:
        date_lst+='-'
print(date_lst)




#Save Review file
final_file='{}_40G&100G.xlsx'.format(date_lst)
if os.path.exists(final_file):
    os.remove(final_file)
    print("Already Removed old Final file")
else:
    print("The Final file file does not exist")

wb2.save(final_file)




#Open Excel File
open_excel = win32com.client.Dispatch('Excel.Application')
open_excel.Visible = True
open_excel.Workbooks.Open("{}/{}".format(os.getcwd(),final_file))


# #Sort By Column  // bug other columns is independent
# open_sheet=open_excel.Sheets("Sheet1")
# open_sheet.Range('A2:A{}'.format(count_cell)).Sort(Key1=open_sheet.Range('A1'), Order1=2, Orientation=1)
# open_sheet.Range('C2:C{}'.format(count_cell)).Sort(Key1=open_sheet.Range('C1'), Order1=1, Orientation=1)
# open_sheet.Range('B2:B{}'.format(count_cell)).Sort(Key1=open_sheet.Range('B1'), Order1=1, Orientation=1)